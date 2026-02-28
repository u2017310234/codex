#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
完整上下文提取器 - 整合所有提取功能
從 Excel 工作簿中提取所有必要的上下文信息，為 LLM 生成代碼做準備
"""

import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from collections import defaultdict
import json
import re


class ExcelContextExtractor:
    """
    完整的 Excel 上下文提取器
    提取所有信息以便 LLM 理解業務邏輯並生成優雅的 Python 代碼
    """
    
    def __init__(self, excel_file: str):
        """
        初始化提取器
        
        Args:
            excel_file: Excel 文件路徑
        """
        self.excel_file = Path(excel_file)
        self.workbook = None
        self.context = {}
        
    def extract_all(self) -> Dict[str, Any]:
        """
        提取所有上下文信息
        
        Returns:
            完整的上下文字典
        """
        print(f"🔍 開始提取: {self.excel_file.name}")
        
        # 加載工作簿
        self.workbook = openpyxl.load_workbook(
            self.excel_file, 
            data_only=False,  # 保留公式
            keep_vba=True     # 保留 VBA
        )
        
        self.context = {
            # 1. 基礎信息
            "metadata": self._extract_metadata(),
            
            # 2. 工作簿結構
            "workbook_structure": self._extract_workbook_structure(),
            
            # 3. 儲存格數值與類型
            "cell_values": self._extract_cell_values(),
            
            # 4. 儲存格公式
            "cell_formulas": self._extract_cell_formulas(),
            
            # 5. 依賴關係
            "dependencies": self._analyze_dependencies(),
            
            # 6. 計算順序
            "calculation_order": self._determine_calculation_order(),
            
            # 7. 數據流分析
            "data_flow": self._analyze_data_flow(),
            
            # 8. VBA 代碼
            "vba_code": self._extract_vba(),
            
            # 9. 重複模式（循環）
            "patterns": self._detect_patterns(),
            
            # 10. 命名範圍
            "named_ranges": self._extract_named_ranges(),
            
            # 11. 表格結構與業務分區
            "table_structure": self._identify_table_structure(),
            
            # 12. 外部連接
            "external_links": self._find_external_links(),
            
            # 13. 條件格式
            "conditional_formatting": self._extract_conditional_formatting(),
            
            # 14. 數據驗證
            "data_validation": self._extract_data_validation(),
        }
        
        self.workbook.close()
        print("✅ 提取完成")
        
        return self.context
    
    def _extract_metadata(self) -> Dict[str, Any]:
        """提取元數據"""
        props = self.workbook.properties
        return {
            "filename": self.excel_file.name,
            "file_size_kb": round(self.excel_file.stat().st_size / 1024, 2),
            "title": props.title,
            "author": props.creator,
            "last_modified_by": props.lastModifiedBy,
            "created": str(props.created) if props.created else None,
            "modified": str(props.modified) if props.modified else None,
            "excel_version": props.version,
        }
    
    def _extract_workbook_structure(self) -> Dict[str, Any]:
        """提取工作簿結構"""
        return {
            "sheet_names": self.workbook.sheetnames,
            "sheet_count": len(self.workbook.sheetnames),
            "active_sheet": self.workbook.active.title if self.workbook.active else None,
            "has_vba": self.workbook.vba_archive is not None,
        }
    
    def _extract_cell_values(self) -> Dict[str, Dict[str, Any]]:
        """
        提取所有儲存格的值與類型
        
        Returns:
            {
                "Sheet1!A1": {
                    "value": 100,
                    "type": "number",
                    "format": "0.00",
                    "is_input": true
                }
            }
        """
        cell_values = {}
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_ref = f"{sheet_name}!{cell.coordinate}"
                        
                        # 判斷類型
                        value = cell.value
                        cell_type = self._get_cell_type(value)
                        
                        # 判斷是輸入值還是公式
                        is_formula = isinstance(value, str) and value.startswith('=')
                        
                        # 转换value为可序列化的类型
                        serializable_value = value
                        if not is_formula:
                            try:
                                # 尝试转换为基本类型
                                if hasattr(value, '__str__') and not isinstance(value, (str, int, float, bool, type(None))):
                                    serializable_value = str(value)
                            except Exception:
                                serializable_value = str(value)
                        else:
                            serializable_value = None
                        
                        cell_values[cell_ref] = {
                            "value": serializable_value,
                            "type": cell_type,
                            "format": cell.number_format,
                            "is_input": not is_formula,
                            "is_formula": is_formula,
                            "row": cell.row,
                            "column": cell.column,
                            "column_letter": cell.column_letter,
                        }
        
        return cell_values
    
    def _get_cell_type(self, value: Any) -> str:
        """判斷儲存格值的類型"""
        if value is None:
            return "empty"
        elif isinstance(value, bool):
            return "boolean"
        elif isinstance(value, (int, float)):
            return "number"
        elif isinstance(value, str):
            if value.startswith('='):
                return "formula"
            return "string"
        else:
            return "other"
    
    def _extract_cell_formulas(self) -> Dict[str, Dict[str, Any]]:
        """
        提取所有公式及其結構化信息
        
        Returns:
            {
                "Sheet1!I31": {
                    "raw_formula": "=LN(...)",
                    "depends_on": ["I24", "I20", ...],
                    "used_functions": ["LN", "EXP", "SQRT"],
                    "complexity": "high"
                }
            }
        """
        formulas = {}
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        cell_ref = f"{sheet_name}!{cell.coordinate}"
                        formula = cell.value
                        
                        formulas[cell_ref] = {
                            "raw_formula": formula,
                            "depends_on": self._extract_cell_references(formula),
                            "used_functions": self._extract_functions(formula),
                            "complexity": self._assess_formula_complexity(formula),
                            "length": len(formula),
                        }
        
        return formulas
    
    def _extract_cell_references(self, formula: str) -> List[str]:
        """從公式中提取儲存格引用"""
        references = []
        
        # 匹配 A1, Sheet1!A1, A1:B10 等
        # 單一儲存格
        pattern1 = r'\b([A-Z]+\d+)\b'
        # 範圍
        pattern2 = r'\b([A-Z]+\d+:[A-Z]+\d+)\b'
        # 跨表引用
        pattern3 = r"('[^']+?'|[A-Z]\w*)!([A-Z]+\d+(?::[A-Z]+\d+)?)"
        
        for match in re.finditer(pattern3, formula):
            references.append(f"{match.group(1)}!{match.group(2)}")
        
        for match in re.finditer(pattern2, formula):
            references.append(match.group(1))
        
        for match in re.finditer(pattern1, formula):
            ref = match.group(1)
            # 避免匹配到函數名（如 A1 在 SUM(A1:A10) 中）
            if ref not in references and not any(ref in r for r in references):
                references.append(ref)
        
        return list(set(references))
    
    def _extract_functions(self, formula: str) -> List[str]:
        """從公式中提取使用的 Excel 函數"""
        functions = re.findall(r'\b([A-Z][A-Z0-9_.]*)\s*\(', formula)
        return list(set(functions))
    
    def _assess_formula_complexity(self, formula: str) -> str:
        """評估公式複雜度"""
        length = len(formula)
        func_count = len(self._extract_functions(formula))
        nesting_level = formula.count('(')
        
        if length > 200 or func_count > 5 or nesting_level > 5:
            return "high"
        elif length > 100 or func_count > 3 or nesting_level > 3:
            return "medium"
        else:
            return "low"
    
    def _analyze_dependencies(self) -> Dict[str, Any]:
        """
        分析儲存格依賴關係
        
        Returns:
            {
                "Sheet1!I31": {
                    "direct_depends": ["I24", "I20", ...],
                    "depended_by": ["I35", "I36"]
                }
            }
        """
        dependencies = defaultdict(lambda: {"direct_depends": [], "depended_by": []})
        formulas = self.context.get("cell_formulas", self._extract_cell_formulas())
        
        for cell_ref, formula_info in formulas.items():
            depends_on = formula_info["depends_on"]
            
            # 規範化引用（加上 sheet 名稱）
            sheet = cell_ref.split('!')[0]
            normalized_deps = []
            for dep in depends_on:
                if '!' not in dep:
                    dep = f"{sheet}!{dep}"
                normalized_deps.append(dep)
            
            dependencies[cell_ref]["direct_depends"] = normalized_deps
            
            # 反向依賴
            for dep in normalized_deps:
                dependencies[dep]["depended_by"].append(cell_ref)
        
        return dict(dependencies)
    
    def _determine_calculation_order(self) -> List[str]:
        """
        確定計算順序（拓撲排序）
        
        Returns:
            按照依賴順序排列的儲存格列表
        """
        dependencies = self.context.get("dependencies", self._analyze_dependencies())
        
        # 簡化版拓撲排序
        visited = set()
        order = []
        
        def visit(cell):
            if cell in visited:
                return
            visited.add(cell)
            
            deps = dependencies.get(cell, {}).get("direct_depends", [])
            for dep in deps:
                visit(dep)
            
            order.append(cell)
        
        # 訪問所有有公式的儲存格
        for cell in dependencies.keys():
            visit(cell)
        
        return order
    
    def _analyze_data_flow(self) -> List[Dict[str, Any]]:
        """
        分析數據流（輸入 → 中間計算 → 輸出）
        
        Returns:
            [
                {"stage": 1, "description": "輸入參數", "cells": [...]},
                {"stage": 2, "description": "中間計算", "cells": [...]},
                ...
            ]
        """
        dependencies = self.context.get("dependencies", self._analyze_dependencies())
        cell_values = self.context.get("cell_values", self._extract_cell_values())
        
        # 分類儲存格
        input_cells = []    # 無依賴的輸入值
        calc_cells = []     # 有依賴的計算儲存格
        output_cells = []   # 無人依賴的輸出值
        
        for cell_ref, cell_info in cell_values.items():
            deps = dependencies.get(cell_ref, {})
            
            if cell_info["is_input"]:
                input_cells.append(cell_ref)
            elif cell_info["is_formula"]:
                if not deps.get("depended_by"):
                    output_cells.append(cell_ref)
                else:
                    calc_cells.append(cell_ref)
        
        return [
            {"stage": 1, "description": "輸入參數", "cells": input_cells, "count": len(input_cells)},
            {"stage": 2, "description": "中間計算", "cells": calc_cells, "count": len(calc_cells)},
            {"stage": 3, "description": "輸出結果", "cells": output_cells, "count": len(output_cells)},
        ]
    
    def _extract_vba(self) -> Dict[str, Any]:
        """
        提取 VBA 代碼
        
        Returns:
            {"has_vba": true, "modules": [...], "note": "..."}
        """
        if not self.workbook.vba_archive:
            return {"has_vba": False}
        
        return {
            "has_vba": True,
            "note": "VBA 代碼已檢測到，詳細分析需要 python-vba 庫",
        }
    
    def _detect_patterns(self) -> List[Dict[str, Any]]:
        """
        檢測重複模式（循環結構）
        
        Returns:
            [
                {
                    "type": "row_loop",
                    "description": "每行使用相同公式模板",
                    "range": "A2:A100",
                    "formula_template": "=B{row}*C{row}",
                    "count": 99
                }
            ]
        """
        patterns = []
        formulas = self.context.get("cell_formulas", self._extract_cell_formulas())
        
        # 按工作表分組
        by_sheet = defaultdict(list)
        for cell_ref, formula_info in formulas.items():
            sheet, cell = cell_ref.split('!')
            by_sheet[sheet].append((cell, formula_info["raw_formula"]))
        
        # 檢測每個工作表的模式
        for sheet, cell_formulas in by_sheet.items():
            # 按列分組
            by_column = defaultdict(list)
            for cell, formula in cell_formulas:
                col = re.match(r'([A-Z]+)', cell).group(1)
                row = int(re.search(r'(\d+)', cell).group(1))
                by_column[col].append((row, formula))
            
            # 檢測列中的重複模式
            for col, formulas_in_col in by_column.items():
                if len(formulas_in_col) >= 3:
                    # 檢查是否是相同模式
                    template = self._extract_formula_template(formulas_in_col)
                    if template:
                        rows = sorted([r for r, _ in formulas_in_col])
                        patterns.append({
                            "type": "column_loop",
                            "sheet": sheet,
                            "column": col,
                            "description": f"列 {col} 中的重複公式模式",
                            "range": f"{sheet}!{col}{rows[0]}:{col}{rows[-1]}",
                            "formula_template": template,
                            "count": len(formulas_in_col),
                            "rows": rows,
                        })
        
        return patterns
    
    def _extract_formula_template(self, formulas: List[Tuple[int, str]]) -> Optional[str]:
        """從一組公式中提取模板"""
        if len(formulas) < 2:
            return None
        
        # 簡化：將行號替換為 {row}
        templates = []
        for row, formula in formulas:
            template = re.sub(r'\b([A-Z]+)' + str(row) + r'\b', r'\1{row}', formula)
            templates.append(template)
        
        # 檢查是否都相同
        if len(set(templates)) == 1:
            return templates[0]
        
        return None
    
    def _extract_named_ranges(self) -> Dict[str, Any]:
        """提取命名範圍"""
        named_ranges = {}
        
        if self.workbook.defined_names:
            for name in self.workbook.defined_names.definedName:
                named_ranges[name.name] = {
                    "refers_to": name.value,
                    "scope": name.localSheetId if name.localSheetId is not None else "workbook",
                }
        
        return named_ranges
    
    def _identify_table_structure(self) -> Dict[str, Any]:
        """識別表格結構與業務分區"""
        # 簡化版：識別有數據的區域
        structure = {}
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            
            if ws.max_row > 0 and ws.max_column > 0:
                structure[sheet_name] = {
                    "data_range": f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}",
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                }
        
        return structure
    
    def _find_external_links(self) -> List[Dict[str, Any]]:
        """查找外部連接"""
        external_links = []
        
        formulas = self.context.get("cell_formulas", self._extract_cell_formulas())
        
        for cell_ref, formula_info in formulas.items():
            formula = formula_info["raw_formula"]
            
            # 檢測外部工作簿引用 [WorkbookName]Sheet!A1
            ext_refs = re.findall(r'\[([^\]]+)\]', formula)
            for ext_ref in ext_refs:
                external_links.append({
                    "type": "external_workbook",
                    "referenced_in": cell_ref,
                    "external_file": ext_ref,
                })
        
        return external_links
    
    def _extract_conditional_formatting(self) -> List[Dict[str, Any]]:
        """提取條件格式"""
        conditional_formatting = []
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            
            if ws.conditional_formatting:
                for cf_range, rules in ws.conditional_formatting._cf_rules.items():
                    for rule in rules:
                        conditional_formatting.append({
                            "sheet": sheet_name,
                            "range": str(cf_range),
                            "type": rule.type if hasattr(rule, 'type') else "unknown",
                        })
        
        return conditional_formatting
    
    def _extract_data_validation(self) -> List[Dict[str, Any]]:
        """提取數據驗證規則"""
        data_validation = []
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            
            if ws.data_validations:
                for dv in ws.data_validations.dataValidation:
                    data_validation.append({
                        "sheet": sheet_name,
                        "range": dv.sqref.ranges[0] if dv.sqref else "unknown",
                        "type": dv.type,
                        "formula1": dv.formula1,
                        "formula2": dv.formula2,
                    })
        
        return data_validation
    
    def save_context(self, output_file: str):
        """將上下文保存為 JSON 文件"""
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.context, f, indent=2, ensure_ascii=False)
        
        print(f"✅ 上下文已保存: {output_path}")


def main():
    """示範用法"""
    # 使用 Margin Call 表格
    excel_file = "/workspaces/RM_Tools/Margin Call.xlsm"
    
    extractor = ExcelContextExtractor(excel_file)
    context = extractor.extract_all()
    
    # 保存上下文
    output_file = "/workspaces/RM_Tools/excel_to_code/output/contexts/margin_call_context.json"
    extractor.save_context(output_file)
    
    # 打印摘要
    print("\n" + "=" * 80)
    print("📊 上下文提取摘要")
    print("=" * 80)
    print(f"檔案名稱: {context['metadata']['filename']}")
    print(f"工作表數: {context['workbook_structure']['sheet_count']}")
    print(f"儲存格數: {len(context['cell_values'])}")
    print(f"公式數: {len(context['cell_formulas'])}")
    print(f"依賴關係: {len(context['dependencies'])}")
    print(f"檢測到的模式: {len(context['patterns'])}")
    print(f"命名範圍: {len(context['named_ranges'])}")
    print(f"VBA 代碼: {'是' if context['vba_code']['has_vba'] else '否'}")
    print("=" * 80)


if __name__ == '__main__':
    main()
